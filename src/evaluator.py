"""
src/evaluator.py
────────────────
Evaluation workflow: loads a manually-labelled Excel file, runs the
classifier, compares Actual_Category vs Predicted_Category, and exports
a colour-coded results workbook with accuracy metrics.

Usage (from project root):
    python3 src/evaluator.py
    python3 src/evaluator.py data/input/evaluation_sample.xlsx
    python3 -m src.evaluator --help
"""

import argparse
import logging
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Ensure project root is importable when run directly ──────────────────────
_project_root = Path(__file__).resolve().parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

from src.loader import load_statement
from src.scorer import score_transaction
from utils.constants import (
    CATEGORY_GST, CATEGORY_NORMAL, CATEGORY_TDS, CATEGORY_POSSIBLE_GST,
    INTERNAL_COLS, TAX_CATEGORY_ORDER,
)
from utils.helpers import normalize_columns

logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────
DEFAULT_EVAL_INPUT  = "data/input/evaluation_sample.xlsx"
DEFAULT_EVAL_OUTPUT = "data/output/evaluation_results.xlsx"
ACTUAL_COL          = "Actual_Category"
PREDICTED_COL       = "Predicted_Category"
CORRECT_COL         = "Correct"
ALL_CATEGORIES      = TAX_CATEGORY_ORDER

# Excel fill colours
_FILL_CORRECT = PatternFill(fill_type="solid", fgColor="FFD9EAD3")  # soft green
_FILL_WRONG   = PatternFill(fill_type="solid", fgColor="FFFCE5CD")  # soft red
_FILL_HEADER  = PatternFill(fill_type="solid", fgColor="FF7030A0")  # purple
_FONT_HEADER  = Font(bold=True, color="FFFFFFFF")


# ── Public API ────────────────────────────────────────────────────────────────

def load_evaluation_file(file_path: Path) -> pd.DataFrame:
    """
    Load the evaluation Excel file using the standard loader.
    Validates that Actual_Category column is present.
    Drops rows where Actual_Category is blank.
    """
    if not file_path.exists():
        sys.exit(f"[ERROR] Evaluation file not found: {file_path}")

    df = load_statement(file_path)

    if ACTUAL_COL not in df.columns:
        sys.exit(
            f"[ERROR] Column '{ACTUAL_COL}' not found in {file_path.name}.\n"
            f"Available columns: {list(df.columns)}\n"
            f"Add an '{ACTUAL_COL}' column with GST / TDS / NORMAL labels."
        )

    # Normalise actual labels
    df[ACTUAL_COL] = (
        df[ACTUAL_COL].astype(str).str.strip().str.upper()
        .replace({"NAN": "", "NONE": "", "NAT": ""})
    )
    unlabelled = (df[ACTUAL_COL] == "").sum()
    if unlabelled:
        logger.warning("Dropping %d rows with missing Actual_Category.", unlabelled)
    df = df[df[ACTUAL_COL] != ""].reset_index(drop=True)

    if df.empty:
        sys.exit(
            f"[ERROR] No labelled rows found in {file_path.name}.\n"
            f"Make sure the '{ACTUAL_COL}' column is filled with "
            f"GST, TDS, or NORMAL for each transaction you want to evaluate."
        )

    logger.info("Evaluation file: %d labelled rows.", len(df))
    return df


def run_classification(df: pd.DataFrame) -> pd.DataFrame:
    """Run the scorer on every row; add Predicted_Category and Confidence."""
    df = df.copy()
    score_results = df.apply(score_transaction, axis=1)
    # ── Map results to new columns ────────────────────────────────────────────
    df["Predicted_Category"]  = [r.category            for r in score_results]
    df["Confidence"]          = [r.confidence          for r in score_results]
    df["Review_Recommended"]  = [r.needs_review        for r in score_results]
    df["Reason"]              = [r.reason              for r in score_results]

    from src.ml_pipeline import append_to_training_data
    append_to_training_data(df, list(score_results))

    # Drop internal alias columns
    df.drop(columns=[c for c in INTERNAL_COLS if c in df.columns], inplace=True)
    return df


def compute_metrics(df: pd.DataFrame) -> dict:
    """
    Compute overall accuracy, per-category breakdown, and confusion summary.
    Adds a Correct column to the DataFrame.
    Returns a metrics dict (includes the augmented df).
    """
    if df.empty:
        return {
            "total": 0, "correct": 0, "accuracy": 0.0,
            "per_category": {}, "confusion": {}, "df": df,
        }

    df = df.copy()
    # Coerce to str first to handle any mixed-dtype edge cases
    df[PREDICTED_COL] = df[PREDICTED_COL].astype(str).str.upper()
    
    # Treat POSSIBLE_GST as correct if actual is GST (sub-category match)
    df[CORRECT_COL] = (df[ACTUAL_COL] == df[PREDICTED_COL]) | (
        (df[ACTUAL_COL] == CATEGORY_GST) & (df[PREDICTED_COL] == CATEGORY_POSSIBLE_GST)
    )

    total    = len(df)
    correct  = int(df[CORRECT_COL].sum())
    accuracy = (correct / total * 100) if total else 0.0

    per_category: dict = {}
    for cat in ALL_CATEGORIES:
        subset = df[df[ACTUAL_COL] == cat]
        if subset.empty:
            continue
        n_correct = int(subset[CORRECT_COL].sum())
        per_category[cat] = {
            "total":    len(subset),
            "correct":  n_correct,
            "wrong":    len(subset) - n_correct,
            "accuracy": (n_correct / len(subset) * 100),
        }

    # Confusion: actual → predicted (only misclassifications)
    confusion: dict = defaultdict(lambda: defaultdict(int))
    for _, row in df[~df[CORRECT_COL]].iterrows():
        confusion[row[ACTUAL_COL]][row[PREDICTED_COL]] += 1

    return {
        "total":        total,
        "correct":      correct,
        "accuracy":     accuracy,
        "per_category": per_category,
        "confusion":    dict(confusion),
        "df":           df,
    }


def print_report(metrics: dict) -> None:
    """Print a formatted evaluation report to stdout."""
    print("\n" + "=" * 55)
    print("  Evaluation Report")
    print("=" * 55)
    print(f"\n  Total labelled  : {metrics['total']:>5,}")
    print(f"  Correct         : {metrics['correct']:>5,}")
    print(f"  Overall Accuracy: {metrics['accuracy']:>5.1f}%")

    print(f"\n  {'Category':<12} {'Total':>6} {'Correct':>8} {'Wrong':>6} {'Accuracy':>9}")
    print(f"  {'-'*12} {'-'*6} {'-'*8} {'-'*6} {'-'*9}")
    for cat, s in metrics["per_category"].items():
        print(f"  {cat:<12} {s['total']:>6,} {s['correct']:>8,} {s['wrong']:>6,} {s['accuracy']:>8.1f}%")

    if metrics["confusion"]:
        print("\n  Confusion Summary (misclassifications only):")
        print(f"  {'Actual':<12} → {'Predicted':<12} {'Count':>6}")
        print(f"  {'-'*12}   {'-'*12} {'-'*6}")
        for actual, preds in sorted(metrics["confusion"].items()):
            for predicted, count in sorted(preds.items(), key=lambda x: -x[1]):
                print(f"  {actual:<12} → {predicted:<12} {count:>6,}")
    else:
        print("\n  ✅ No misclassifications found!")
    print("\n" + "=" * 55)


def export_results(metrics: dict, output_path: Path) -> None:
    """
    Write a two-sheet Excel workbook:
      - Results: all rows, colour-coded (green=correct, red=wrong)
      - Summary: accuracy metrics and confusion table
    """
    df = metrics["df"]
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Reorder: original cols first, then eval cols
    eval_cols  = [c for c in [ACTUAL_COL, PREDICTED_COL, "Confidence", CORRECT_COL] if c in df.columns]
    other_cols = [c for c in df.columns if c not in eval_cols]
    df = df[other_cols + eval_cols]

    # Build summary DataFrame
    summary_rows = [
        {"Metric": "Total Evaluated",  "Value": metrics["total"]},
        {"Metric": "Correct",          "Value": metrics["correct"]},
        {"Metric": "Overall Accuracy", "Value": f"{metrics['accuracy']:.1f}%"},
        {"Metric": "", "Value": ""},
    ]
    for cat, s in metrics["per_category"].items():
        summary_rows += [
            {"Metric": f"{cat} — Total",    "Value": s["total"]},
            {"Metric": f"{cat} — Correct",  "Value": s["correct"]},
            {"Metric": f"{cat} — Accuracy", "Value": f"{s['accuracy']:.1f}%"},
            {"Metric": "", "Value": ""},
        ]
    if metrics["confusion"]:
        summary_rows.append({"Metric": "─── Confusion ───", "Value": ""})
        for actual, preds in sorted(metrics["confusion"].items()):
            for predicted, count in sorted(preds.items(), key=lambda x: -x[1]):
                summary_rows.append({"Metric": f"{actual} → {predicted}", "Value": count})

    summary_df = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Results", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    _format_workbook(output_path, CORRECT_COL)
    logger.info("Evaluation results written → %s", output_path)


# ── Private helpers ───────────────────────────────────────────────────────────

def _format_workbook(path: Path, correct_col_name: str) -> None:
    try:
        wb = load_workbook(path)

        # ── Results sheet ──────────────────────────────────────────────────────
        ws = wb["Results"]
        # Find Correct column index
        correct_idx = next(
            (i for i, c in enumerate(ws[1], 1) if c.value == correct_col_name),
            None,
        )
        for cell in ws[1]:
            cell.fill = _FILL_HEADER
            cell.font = _FONT_HEADER
            cell.alignment = Alignment(horizontal="center")

        if correct_idx:
            for row in ws.iter_rows(min_row=2):
                fill = _FILL_CORRECT if row[correct_idx - 1].value is True else _FILL_WRONG
                for cell in row:
                    cell.fill = fill

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

        # ── Summary sheet ──────────────────────────────────────────────────────
        ws2 = wb["Summary"]
        for cell in ws2[1]:
            cell.fill = _FILL_HEADER
            cell.font = _FONT_HEADER
        for col_idx, col_cells in enumerate(ws2.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        wb.save(path)
    except Exception as exc:
        logger.warning("Could not format evaluation results: %s", exc)


# ── CLI ───────────────────────────────────────────────────────────────────────

def _resolve(p: str) -> Path:
    path = Path(p)
    return path if path.is_absolute() else (_project_root / path).resolve()


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="evaluator",
        description="Evaluate the GST/TDS classifier against manually labelled data.",
    )
    parser.add_argument(
        "input_file", nargs="?", default=DEFAULT_EVAL_INPUT,
        help=f"Evaluation xlsx with Actual_Category column (default: {DEFAULT_EVAL_INPUT})",
    )
    parser.add_argument(
        "--output", default=DEFAULT_EVAL_OUTPUT,
        help=f"Output path for results workbook (default: {DEFAULT_EVAL_OUTPUT})",
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(level=level,
                        format="%(asctime)s [%(levelname)-8s] %(name)s — %(message)s",
                        datefmt="%H:%M:%S")

    input_path  = _resolve(args.input_file)
    output_path = _resolve(args.output)

    print("=" * 55)
    print("  GST & TDS Classifier — Evaluation Mode")
    print("=" * 55)

    print(f"\n📂 Loading: {input_path.name}")
    df = load_evaluation_file(input_path)
    print(f"   {len(df):,} labelled transactions loaded.")

    print("\n⚙️  Running classifier ...")
    df = run_classification(df)

    print("📊 Computing metrics ...")
    metrics = compute_metrics(df)

    print_report(metrics)

    export_results(metrics, output_path)
    print(f"\n✅ Results saved → {output_path}")


if __name__ == "__main__":
    main()
