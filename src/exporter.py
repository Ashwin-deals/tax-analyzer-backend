"""
src/exporter.py
───────────────
Writes per-category Excel files and a classification summary workbook.

Exports the final TAX_CATEGORY files: GST, POSSIBLE_GST, TDS, NORMAL.
Summary includes review counts and confidence breakdown.
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.constants import (
    CATEGORY_COLOURS, CATEGORY_GST, CATEGORY_NORMAL,
    CATEGORY_TDS, CATEGORY_POSSIBLE_GST,
    DEFAULT_OUTPUT_DIR, OUTPUT_FILENAMES, SUMMARY_FILENAME, TAX_CATEGORY_ORDER,
)
from utils.helpers import build_summary

logger = logging.getLogger(__name__)


def export_data(
    data_dict: dict[str, pd.DataFrame],
    output_folder: str | Path = DEFAULT_OUTPUT_DIR,
    include_summary: bool = True,
) -> None:
    """
    Export per-category DataFrames to individual formatted Excel files.

    Parameters
    ----------
    data_dict : dict[str, pd.DataFrame]
        Keys: 'GST', 'POSSIBLE_GST', 'TDS', 'NORMAL'.
    output_folder : str | Path
        Destination directory (created if missing).
    include_summary : bool
        If True, also writes classification_summary.xlsx.
    """
    out_dir = Path(output_folder)
    out_dir.mkdir(parents=True, exist_ok=True)

    exported: list[tuple[str, int]] = []

    for category in TAX_CATEGORY_ORDER:
        df = data_dict.get(category, pd.DataFrame())
        filename = OUTPUT_FILENAMES[category]
        dest = out_dir / filename
        _write_sheet(df, dest, category)
        logger.info("Exported %d %s rows → %s", len(df), category, dest)
        exported.append((filename, len(df)))

    if include_summary:
        all_rows = [df for df in data_dict.values() if not df.empty]
        if all_rows:
            combined = pd.concat(all_rows, ignore_index=True)
            summary_df = build_summary(combined)
            summary_path = out_dir / SUMMARY_FILENAME
            _write_sheet(summary_df, summary_path, "SUMMARY")
            logger.info("Summary written → %s", summary_path)

    print(f"\n✅ Export complete → {out_dir.resolve()}")
    for filename, count in exported:
        print(f"   • {filename:<40} ({count:>4} rows)")
    if include_summary:
        print(f"   • {SUMMARY_FILENAME}")


# ── Private helpers ───────────────────────────────────────────────────────────

def _write_sheet(df: pd.DataFrame, dest: Path, category: str) -> None:
    """Write DataFrame to Excel then apply header formatting."""
    df.to_excel(dest, index=False, engine="openpyxl")
    _format_worksheet(dest, category)


def _format_worksheet(path: Path, category: str) -> None:
    """Apply colour-coded headers and auto-width columns (in-place)."""
    try:
        wb = load_workbook(path)
        ws = wb.active

        colour = CATEGORY_COLOURS.get(category, "FF4472C4")
        fill   = PatternFill(fill_type="solid", fgColor=colour)
        font   = Font(bold=True, color="FFFFFFFF")

        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

        wb.save(path)
    except Exception as exc:
        logger.warning("Could not format %s: %s", path.name, exc)
